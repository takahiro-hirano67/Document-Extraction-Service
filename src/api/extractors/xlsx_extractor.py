# src/api/extractors/xlsx_extractor.py

"""Excelファイル(.xlsx)からテキストを抽出するモジュール

セルのテキスト（openpyxl）と図形・テキストボックスのテキスト（lxml）を
それぞれ抽出し、結合して返す。

## 実装の要点

- xlsx は ZIP 形式であり、図形テキストは DrawingML (xl/drawings/drawing*.xml) に格納されている
- セル抽出には openpyxl、図形テキスト抽出と drawing→シート名マッピングには lxml を使用する
- drawing ファイルとシートの対応は xl/_rels/workbook.xml.rels と
  xl/worksheets/_rels/sheet*.xml.rels の 2 段階リレーションシップを辿って解決する

## 備考

- 取得できないもの: ワードアート、グループ化図形の一部、グラフ凡例テキスト
- スキャン画像が貼り付けられたセルのテキストは OCR なしでは抽出不可
"""

# 型チェック時以外は型ヒントを文字列として認識させる
from __future__ import annotations

import re
import zipfile
from datetime import datetime
from typing import TYPE_CHECKING

from lxml import etree
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell

# 型チェック時のみインポートするブロック
if TYPE_CHECKING:
    import io

    from openpyxl.cell.cell import Cell
    from openpyxl.workbook.workbook import Workbook

# ==========================================
# 定数
# ==========================================

# DrawingML 名前空間
_NS_DRAWING_MAIN: str = "http://schemas.openxmlformats.org/drawingml/2006/main"
_NS_SPREADSHEET_DRAWING: str = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
_NS_RELATIONSHIPS: str = "http://schemas.openxmlformats.org/package/2006/relationships"

# Excel の日付シリアル値として妥当な範囲（1990-01-01 〜 2100-12-31 相当）
_EXCEL_DATE_SERIAL_MIN: int = 32874
_EXCEL_DATE_SERIAL_MAX: int = 73050
_EXCEL_EPOCH: datetime = datetime(1899, 12, 30)

# data_only=True でも解決されずにセルに残る Excel のエラー値
_EXCEL_ERROR_VALUES: frozenset[str] = frozenset(
    {
        "#VALUE!",
        "#REF!",
        "#DIV/0!",
        "#NAME?",
        "#N/A",
        "#NULL!",
        "#NUM!",
    }
)

# セル内の読点（全角カンマ）: 列区切りカンマとの混同を防ぐため半角カンマに先立って置換する
_COMMA_REPLACEMENT: str = "、"


# ==========================================
# プライベート関数: セル値の正規化
# ==========================================


def _serial_to_date_string(serial_value: int) -> str | None:
    """Excel の日付シリアル値を YYYY-MM-DD 文字列に変換する。

    変換できない場合は None を返す。
    """
    try:
        date_value = datetime.fromordinal(_EXCEL_EPOCH.toordinal() + serial_value)
        return date_value.strftime("%Y-%m-%d")
    except (ValueError, OverflowError):
        return None


def _normalize_cell_value(cell: Cell) -> str:
    """セル値を文字列に正規化する。

    - Excel エラー値 (#VALUE! 等): 空文字列に変換
    - datetime 型: ISO 形式（YYYY-MM-DD）に変換
    - 日付シリアル値と推定される数値: ISO 形式に変換
    - セル内改行: 半角スペースに置換
    - 桁区切りカンマ（数字,数字）: 除去
    - 半角カンマ: 全角読点（、）に置換
    - その他: str() で文字列化
    """
    value = cell.value

    # Excel エラー値はノイズになるため除去する
    if isinstance(value, str) and value.strip() in _EXCEL_ERROR_VALUES:
        return ""

    # openpyxl が datetime として返した場合
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")

    # data_only=True でもシリアル値（int/float）で返る場合への対処
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        int_value = int(value)
        if _EXCEL_DATE_SERIAL_MIN <= int_value <= _EXCEL_DATE_SERIAL_MAX:
            converted = _serial_to_date_string(int_value)
            if converted is not None:
                return converted

    normalized: str = str(value).replace("\n", " ").strip()

    # 桁区切りカンマを除去する（例: 9,200 → 9200）
    # 全角読点への置換より先に処理することで、数値間のカンマのみを対象にできる
    normalized = re.sub(r"(?<=\d),(?=\d)", "", normalized)

    # 半角カンマを全角読点に置換する（列区切りカンマとの混同を防ぐ）
    normalized = normalized.replace(",", _COMMA_REPLACEMENT)

    return normalized


# ==========================================
# プライベート関数: 空行の圧縮
# ==========================================


def _collapse_blank_lines(lines: list[str]) -> list[str]:
    """連続する空行を 1 行にまとめる。"""
    collapsed: list[str] = []
    previous_was_blank: bool = False

    for line in lines:
        is_blank: bool = line.strip() == ""
        if is_blank and previous_was_blank:
            continue
        collapsed.append(line)
        previous_was_blank = is_blank

    return collapsed


# ==========================================
# プライベート関数: drawing → シート名 マッピング
# ==========================================


def _build_drawing_to_sheet_map(zip_file: zipfile.ZipFile) -> dict[str, str]:
    """xl/workbook.xml.rels と xl/worksheets/_rels/ を辿り、drawing ファイルパス → シート名 のマッピングを返す。

    解決手順:
        1. xl/workbook.xml.rels から rId → sheet*.xml パスを取得
        2. workbook.xml の r:id と照合してシートパス → シート名 を構築
        3. xl/worksheets/_rels/sheet*.xml.rels から drawing パス → シート名 を構築
    """
    drawing_to_sheet: dict[str, str] = {}

    ns_workbook: str = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    ns_r_id: str = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"

    # ── Step 1: workbook.xml を取得 ──────────────────────────────
    workbook_path: str = "xl/workbook.xml"
    if workbook_path not in zip_file.namelist():
        return drawing_to_sheet

    with zip_file.open(workbook_path) as workbook_file:
        workbook_tree = etree.parse(workbook_file)

    # ── Step 2: workbook.xml.rels から rId → sheet*.xml パスを取得 ──
    workbook_rels_path: str = "xl/_rels/workbook.xml.rels"
    if workbook_rels_path not in zip_file.namelist():
        return drawing_to_sheet

    with zip_file.open(workbook_rels_path) as rels_file:
        rels_tree = etree.parse(rels_file)

    rid_to_sheet_path: dict[str, str] = {}
    for rel in rels_tree.findall(f"{{{_NS_RELATIONSHIPS}}}Relationship"):
        rel_type: str = rel.get("Type", "")
        if "worksheet" not in rel_type:
            continue
        rid: str | None = rel.get("Id")
        target: str | None = rel.get("Target")
        if rid and target:
            # Target は "worksheets/sheet1.xml" のような相対パス
            rid_to_sheet_path[rid] = f"xl/{target}"

    # ── Step 3: workbook.xml の r:id と突合してシートパス → シート名 を構築 ──
    sheet_path_to_name: dict[str, str] = {}
    for sheet_element in workbook_tree.findall(f".//{{{ns_workbook}}}sheet"):
        r_id: str | None = sheet_element.get(ns_r_id)
        sheet_name: str | None = sheet_element.get("name")
        if r_id and sheet_name and r_id in rid_to_sheet_path:
            sheet_path: str = rid_to_sheet_path[r_id]
            sheet_path_to_name[sheet_path] = sheet_name

    # ── Step 4: sheet*.xml.rels から drawing パス → シート名 を構築 ──
    for sheet_path, sheet_name in sheet_path_to_name.items():
        # xl/worksheets/sheet1.xml → xl/worksheets/_rels/sheet1.xml.rels
        sheet_filename: str = sheet_path.split("/")[-1]
        sheet_rels_path: str = f"xl/worksheets/_rels/{sheet_filename}.rels"

        if sheet_rels_path not in zip_file.namelist():
            continue

        with zip_file.open(sheet_rels_path) as sheet_rels_file:
            sheet_rels_tree = etree.parse(sheet_rels_file)

        for rel in sheet_rels_tree.findall(f"{{{_NS_RELATIONSHIPS}}}Relationship"):
            rel_type = rel.get("Type", "")
            if "drawing" not in rel_type:
                continue
            target = rel.get("Target")
            if target:
                # Target は "../drawings/drawing1.xml" のような相対パス
                drawing_path: str = "xl/" + target.replace("../", "")
                drawing_to_sheet[drawing_path] = sheet_name

    return drawing_to_sheet


# ==========================================
# プライベート関数: セル・図形テキストの抽出
# ==========================================


def _extract_cell_lines(work_book: Workbook) -> list[str]:
    """Openpyxl を使って全シートのセル値を行テキストのリストとして返す。

    - MergedCell（結合セルの左上以外）はスキップする
    - 各セルの値はカンマ区切りで結合し、列の独立性を示す
    - 空行は空文字列として追加し、呼び出し元で連続空行の圧縮を行う

    """
    cell_lines: list[str] = []

    for sheet in work_book.worksheets:
        cell_lines.append(f"\n=== シート: {sheet.title} ===\n")

        for row in sheet.iter_rows():
            cell_texts: list[str] = []

            for cell in row:
                if isinstance(cell, MergedCell):
                    # 結合セルの左上以外はスキップ（値は左上セルに集約されている）
                    continue
                if cell.value is None:
                    continue

                normalized_value: str = _normalize_cell_value(cell)
                if normalized_value:
                    cell_texts.append(normalized_value)

            # セルをカンマ区切りで結合し、列の独立性をLLMに示す
            row_text: str = ", ".join(cell_texts)
            # 空行も追加する（連続空行の圧縮は _collapse_blank_lines で行う）
            cell_lines.append(row_text)

    return cell_lines


def _extract_shape_lines(
    file_stream: io.BytesIO,
    drawing_to_sheet: dict[str, str],
) -> list[str]:
    """Lxml を使って xlsx 内の図形・テキストボックスのテキストを返す。

    - drawing_to_sheet マップを使い、各図形テキストにシート名を付与する
    - 図形テキストが 1 件も存在しないシートのヘッダーは出力しない
    """
    shape_lines: list[str] = []
    current_sheet_name: str = ""

    with zipfile.ZipFile(file_stream) as zip_file:
        drawing_entry_names: list[str] = sorted(
            [entry_name for entry_name in zip_file.namelist() if re.search(r"xl/drawings/drawing\d+\.xml$", entry_name)]
        )

        for drawing_entry_name in drawing_entry_names:
            sheet_name: str = drawing_to_sheet.get(drawing_entry_name, "不明")

            with zip_file.open(drawing_entry_name) as drawing_file:
                drawing_tree = etree.parse(drawing_file)

            shape_elements = drawing_tree.findall(f".//{{{_NS_SPREADSHEET_DRAWING}}}sp")

            # このdrawingファイルから抽出できる図形テキストを先に収集する
            # テキストが存在しない場合はシートヘッダーごと出力しない
            texts_in_drawing: list[str] = []
            for shape_element in shape_elements:
                text_run_elements = shape_element.findall(f".//{{{_NS_DRAWING_MAIN}}}t")
                shape_text: str = "".join(text_run.text or "" for text_run in text_run_elements).strip()
                if shape_text:
                    texts_in_drawing.append(shape_text)

            if not texts_in_drawing:
                # 図形テキストなし: このdrawingはスキップ（ヘッダーも出力しない）
                continue

            # シートが切り替わったらヘッダーを挿入
            if sheet_name != current_sheet_name:
                if shape_lines:
                    shape_lines.append("")
                shape_lines.append(f"\n=== シート: {sheet_name} ===\n")
                current_sheet_name = sheet_name

            shape_lines.extend(texts_in_drawing)

    return shape_lines


# ==========================================
# パブリック関数
# ==========================================


def extract_from_xlsx(file_stream: io.BytesIO) -> str:
    """Excel ファイルからテキストを抽出する。

    Args:
        file_stream: xlsx ファイルの BytesIO ストリーム

    Returns:
        抽出されたテキスト全体を改行区切りで結合した文字列

    """
    file_stream.seek(0)
    work_book: Workbook = load_workbook(file_stream, data_only=True)
    cell_lines: list[str] = _extract_cell_lines(work_book)

    # drawing → シート名 のマッピングを構築
    file_stream.seek(0)
    with zipfile.ZipFile(file_stream) as zip_file:
        drawing_to_sheet: dict[str, str] = _build_drawing_to_sheet_map(zip_file)

    file_stream.seek(0)
    shape_lines: list[str] = _extract_shape_lines(file_stream, drawing_to_sheet)

    all_lines: list[str] = []

    if cell_lines:
        all_lines.append("\n== セル内のテキスト ==\n")
        all_lines.extend(cell_lines)

    if shape_lines:
        if all_lines:
            all_lines.append("")
        all_lines.append("\n== 図形やテキストボックス内のテキスト ==\n")
        all_lines.extend(shape_lines)

    collapsed_lines: list[str] = _collapse_blank_lines(all_lines)
    result_text: str = "\n".join(collapsed_lines)
    result_text = re.sub(r"\n{3,}", "\n\n", result_text).strip()
    result_prefix = "以下の内容は、XLSX形式のファイルから機械的にテキストを抽出したものです。\n\n"
    return result_prefix + result_text
